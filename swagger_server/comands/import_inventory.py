# scripts/import_inventory.py

import sys
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError

from swagger_server.resources.databases.postgresql import PostgreSQLClient


COLUMN_MAPPING = {
    "CODIGO": "code",
    "Codigo": "code",
    "MODELO": "model",
    "PRODUCTO": "product",
    "Unid. De medida": "unit",
    "COSTO BASE": "base_price",
    "% de ganancia": "profit_margin",
    "% de ganancia en $": "profit_margin_dollar",
    "Precio de venta": "price",
    "Proveedor": "provider",
    "Observacion": "description",
    "STOCK": "stock",
}


def load_sheet(ws):
    data = list(ws.values)

    if not data:
        return pd.DataFrame()

    headers = data[0]
    rows = data[1:]

    return pd.DataFrame(rows, columns=headers)


def clean_dataframe(df):

    df = df.rename(columns=COLUMN_MAPPING)

    # Solo conservar columnas existentes en la tabla
    columns = [
        "code",
        "model",
        "product",
        "unit",
        "base_price",
        "profit_margin",
        "profit_margin_dollar",
        "price",
        "provider",
        "description",
        "stock",
    ]

    df = df[[c for c in columns if c in df.columns]]

    # Eliminar filas completamente vacías
    df = df.dropna(how="all")

    # Eliminar filas basura (FILTER, REF!, etc.)
    if "code" in df.columns:
        df = df[
            ~df["code"]
            .astype(str)
            .str.contains("FILTER|#REF!|DUMMYFUNCTION", case=False, na=False)
        ]

    # Convertir NaN -> None
    df = df.replace({np.nan: None})

    numeric_cols = [
        "base_price",
        "profit_margin",
        "profit_margin_dollar",
        "price",
        "stock",
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.replace({np.nan: None})

    return df


def main(filepath, db_name="POSTGRESQL"):

    client = PostgreSQLClient(db_name)
    engine = client.engine

    workbook = load_workbook(
        filename=filepath,
        data_only=True
    )

    total = 0

    for sheet_name in workbook.sheetnames:

        print("=" * 80)
        print(f"Procesando hoja: {sheet_name}")

        try:

            ws = workbook[sheet_name]

            df = load_sheet(ws)
            df = clean_dataframe(df)

            if df.empty:
                print("Sin registros.")
                continue

            print(f"Registros a insertar: {len(df)}")

            df.to_sql(
                name="technical_equipment",
                schema="technical",
                con=engine,
                if_exists="append",
                index=False,
                method="multi",
                chunksize=500,
            )

            total += len(df)

            print(f"OK -> {len(df)} registros")

        except SQLAlchemyError as e:
            print(f"Error SQL en hoja {sheet_name}")
            print(e)

        except Exception as e:
            print(f"Error en hoja {sheet_name}")
            print(e)

    print("=" * 80)
    print(f"Importación finalizada. Total insertados: {total}")


if __name__ == "__main__":

    filepath = sys.argv[1]
    db_name = sys.argv[2] if len(sys.argv) > 2 else "POSTGRESQL"

    main(filepath, db_name)

    # python -m scripts.import_inventory "precios.xlsx"